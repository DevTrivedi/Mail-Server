#author "Dev Trivedi & Monil Shah"

from tkinter import *
import smtplib
import os
import getpass
import xlrd
import datetime
import openpyxl
import re
import socket
import coloredlogs,logging

from easygui import passwordbox
from tkinter import filedialog
from tkinter import *
from string import Template
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders


# Create a logger object.
logger = logging.getLogger(__name__)
logger.warning("\t\t\t\tWelcome to DM Mail Server\t\t\t\t")
print()
logger.warning("Version: "+"1.0")
print()
logger.warning("Description: "+"Useful mail server to help you with your requirements of sending mails in daily life. ")
print()
logger.warning("Author: "+"Dev Trivedi & Monil Shah")
print()
print()

status = False

while(status == False):

    email =input("Enter Sender's Email Address : ")             # example@domain.com    
    isvalid=False
    zv=email.split('@')
    zv[1][:zv[1].index(".com")]
    if(len(zv)):
        print()
        print("Valid Email Id")
        print()
        isValid=True
        
    n=100    

    for i in range(1,n):
        
        if (zv == None): 
            print()
            print("Invalid Email-id !! ")
            print()
            print("Please enter the email in format: example@domain.com ")
            print()
            email =input("Enter Sender's Email Address : ")

# set up the SMTP server
    try:
        password = passwordbox('Enter Password : ')         # In the prompt box enter the password of your email id.
        s = smtplib.SMTP('smtp.gmail.com',587)
        s.starttls()
        print()
        print("Please wait for a while...")
        print()
        s.login(email, password)
        print("Login Successful !!")
        print()

# Enter the body message of your mail.
        print("Enter your message and press 'Esc' Key to end the message")
        print()
        message = '  '
        while ((message[-2]) != '$'):
            message+=(input())
            message+=('\n')
        message=message[2:-2]

# Enter the subject of your mail.
        subject=input('Enter Subject of your mail : ')
        print()    
        print("Please select the excel file (.xlsx) you want to use with this server: ")
        print()
        print('Select File') # Select the excel file in which you have provided the email id with the attachment of the file and the date properly.
        root = Tk()
        root.excelfilename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("Excel Files","*.xlsx*"),("Excel Files","*.xlsx*")))
        print()
        root.destroy()
        wbxl = openpyxl.load_workbook(root.excelfilename)
        
        sheetname = input("Enter the spread-sheet name from the Excel file selected: ")
        wbxl = openpyxl.load_workbook(root.excelfilename)
        sheetxl=wbxl[sheetname]            
        wb = xlrd.open_workbook(root.excelfilename)        
        sheet = wb.sheet_by_name(sheetname)
        today=datetime.datetime.now().strftime("%d")

        for i in range(1,sheet.nrows):
            To = sheet.cell_value(i,0)
            filename = sheet.cell_value(i,1)
            dt=sheet.cell_value(i,2).split("/")[0]
            
            if((int(today)==(int(dt) if len(dt)!=0 else 0))and sheet.cell_value(i,3)!='Yes'):
                try:
                    attachment = open(filename,"rb")
                    nofile=False
                except:
                    logger.warning("Please check your excel file, it has some rows without any attachment path...")
                    nofile=True
                msg = MIMEMultipart()

                #Sender Address
                # setup the parameters of the message
                msg['From']=email
                msg['Subject']=subject
                msg.attach(MIMEText(message, 'plain'))
                p = MIMEBase('application','octet-stream')
                # To change the payload into encoded form
                
                if(nofile==False):
                    p.set_payload((attachment).read())
                # encode into base64
                    encoders.encode_base64(p)
                    p.add_header('Content-Disposition', "attachment; filename= %s" %filename.split('/')[-1])
                # attach the instance 'p' to instance 'msg'
                    msg.attach(p)
                del filename
                
                s.sendmail(email,To,msg.as_string())
                sheetxl.cell(row=i+1, column=4).value="Yes"
                wbxl.save(root.excelfilename)
                del msg

            status = True

        # Terminate the SMTP session and close the connection
        #s.close()
        s.quit

    except smtplib.SMTPAuthenticationError:
        print()
        logger.warning("Error occured while signing into your account. ")
        print()
        logger.warning("Possible reasons can be: ")
        print()
        logger.warning("1) You entered wrong Email/Password.")
        print()
        logger.warning("2) Make sure 2-step verification is disabled in your Gmail account. ")
        print()
        logger.warning("3) Please make sure you have allowed the less secure apps option in 'My account' settings of Gmail. ")
        print()
        logger.warning("Please try again...")
        print()
        
    except socket.gaierror:
        print()
        logger.warning("Oops!! Looks like you are not connected to the internet or have proxy enabled. ")
        print()
        logger.warning("Please check your internet connection or system proxy and try again. ")
        print()

    except AttributeError:
        print()
        print()
        logger.warning("Please enter the password to continue...")
        print()

    except FileNotFoundError:
       print()
       logger.warning("Please select the excel file to use this server...")
       print()

    except ValueError:
        print()
        logger.warning("Please enter the email in format: example@domain.com ")
        print()
